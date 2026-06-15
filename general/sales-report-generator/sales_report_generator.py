"""
Automated Sales Report Generator
Produces a 4-sheet professional Excel report from raw sales data.

Usage:
    python sales_report_generator.py
    → generates Sales_Report_Q1_2025.xlsx

Requirements:
    pip install openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from collections import defaultdict

# ============================================================
# CONFIGURATION — swap this with your own data source
# ============================================================
SALES_DATA = [
    # Product, Category, Region, Units, Price, Cost
    ["Wireless Mouse", "Accessories", "North", 145, 24.99, 12.50],
    ["Mechanical Keyboard", "Accessories", "North", 98, 89.99, 45.00],
    ["USB-C Hub", "Accessories", "South", 210, 34.50, 17.25],
    ["27\" Monitor", "Displays", "East", 45, 299.99, 180.00],
    ["24\" Monitor", "Displays", "West", 67, 189.99, 110.00],
    ["Webcam 1080p", "Peripherals", "North", 167, 54.99, 27.50],
    ["Laptop Stand", "Accessories", "South", 320, 39.99, 15.00],
    ["Desk Lamp LED", "Office", "East", 275, 29.99, 10.00],
    ["Standing Desk", "Office", "West", 22, 499.99, 250.00],
    ["Noise Canceling Phones", "Audio", "North", 89, 149.99, 75.00],
    ["Bluetooth Speaker", "Audio", "South", 134, 79.99, 40.00],
    ["HDMI Cable 6ft", "Accessories", "East", 450, 12.99, 4.00],
    ["Wireless Charger", "Accessories", "West", 189, 19.99, 8.00],
    ["Ergonomic Chair", "Office", "North", 15, 899.99, 450.00],
]

# ============================================================
# STYLES
# ============================================================
DARK_BLUE, MED_BLUE, LIGHT_BLUE = "1F4E79", "2E75B6", "D6E4F0"
GREEN_HEADER, LIGHT_GREEN, WHITE = "375623", "E2EFDA", "FFFFFF"

hdr_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
hdr_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
alt_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
curr_fmt, num_fmt, pct_fmt = '$#,##0.00', '#,##0', '0.0%'

# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = Workbook()

# ----- SHEET 1: Detailed Sales Data -----
ws = wb.active
ws.title = "Sales Data"

headers = ["Product", "Category", "Region", "Units Sold",
           "Unit Price", "Unit Cost", "Revenue", "Profit", "Margin %"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, hdr_align, thin_border

for r, row in enumerate(SALES_DATA, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name="Calibri", size=10)
        cell.border = thin_border
        if c >= 4:
            cell.alignment = Alignment(horizontal="right")
        if r % 2 == 0:
            cell.fill = alt_fill

    # Formulas
    ws.cell(row=r, column=7, value=f"=D{r}*E{r}")
    ws.cell(row=r, column=8, value=f"=G{r}-(D{r}*F{r})")
    ws.cell(row=r, column=9, value=f"=H{r}/G{r}")
    ws.cell(row=r, column=5).number_format = curr_fmt
    ws.cell(row=r, column=6).number_format = curr_fmt
    ws.cell(row=r, column=7).number_format = curr_fmt
    ws.cell(row=r, column=8).number_format = curr_fmt
    ws.cell(row=r, column=9).number_format = pct_fmt

# Totals row
tr = len(SALES_DATA) + 2
ws.merge_cells(f"A{tr}:C{tr}")
tc = ws.cell(row=tr, column=1, value="TOTALS")
tc.font = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
tc.alignment = Alignment(horizontal="right")
tc.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
for col in [4, 7, 8]:
    cell = ws.cell(row=tr, column=col)
    cell.value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{tr - 1})"
    cell.font = Font(name="Calibri", size=11, bold=True)
    cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    cell.border = thin_border
    cell.number_format = curr_fmt if col >= 7 else num_fmt

# Conditional formatting + freeze + auto-filter
ws.conditional_formatting.add(
    f"I2:I{tr - 1}",
    ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="num", end_value=0.6, end_color="63BE7B"))
ws.conditional_formatting.add(
    f"D2:D{tr - 1}",
    DataBarRule(start_type="min", end_type="max", color=MED_BLUE, showValue=True))
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{tr - 1}"

# Column widths
for i, w in enumerate([26, 16, 10, 12, 14, 14, 14, 14, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ----- SHEET 2: Category Summary -----
ws_cat = wb.create_sheet("Category Summary")

# Aggregate by category
cat_data = defaultdict(lambda: {"units": 0, "revenue": 0.0, "profit": 0.0})
for row in SALES_DATA:
    cat, units, price, cost = row[1], row[3], row[4], row[5]
    rev = units * price
    cat_data[cat]["units"] += units
    cat_data[cat]["revenue"] += rev
    cat_data[cat]["profit"] += rev - (units * cost)

sorted_cats = sorted(cat_data.items(), key=lambda x: x[1]["revenue"], reverse=True)

for c, h in enumerate(["Category", "Total Units", "Total Revenue", "Total Profit", "Margin %"], 1):
    cell = ws_cat.cell(row=1, column=c, value=h)
    cell.fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type="solid")
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.alignment, cell.border = hdr_align, thin_border

for r, (cat, vals) in enumerate(sorted_cats, 2):
    ws_cat.cell(row=r, column=1, value=cat)
    ws_cat.cell(row=r, column=2, value=vals["units"])
    ws_cat.cell(row=r, column=3, value=round(vals["revenue"], 2))
    ws_cat.cell(row=r, column=4, value=round(vals["profit"], 2))
    ws_cat.cell(row=r, column=5, value=f"=D{r}/C{r}")
    for c in range(1, 6):
        cell = ws_cat.cell(row=r, column=c)
        cell.font = Font(name="Calibri", size=10)
        cell.border = thin_border
        if c >= 2:
            cell.alignment = Alignment(horizontal="right")
        if r % 2 == 0:
            cell.fill = alt_fill
    ws_cat.cell(row=r, column=3).number_format = curr_fmt
    ws_cat.cell(row=r, column=4).number_format = curr_fmt
    ws_cat.cell(row=r, column=5).number_format = pct_fmt

# Bar chart
bar = BarChart()
bar.type = "col"
bar.style = 10
bar.title = "Revenue by Product Category"
bar.width, bar.height = 22, 14
bar.add_data(Reference(ws_cat, min_col=3, min_row=1, max_row=len(sorted_cats) + 1),
             titles_from_data=True)
bar.set_categories(Reference(ws_cat, min_col=1, min_row=2, max_row=len(sorted_cats) + 1))
chart_colors = ["2E75B6", "ED7D31", "A5A5A5", "FFC000", "4472C4", "70AD47"]
for i in range(len(sorted_cats)):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = chart_colors[i % 6]
    bar.series[0].data_points.append(pt)
ws_cat.add_chart(bar, "G2")
ws_cat.conditional_formatting.add(
    f"C2:C{len(sorted_cats) + 1}",
    DataBarRule(start_type="min", end_type="max", color=MED_BLUE, showValue=True))
for i, w in enumerate([20, 14, 18, 16, 12], 1):
    ws_cat.column_dimensions[get_column_letter(i)].width = w

# ----- SHEET 3: Regional Breakdown -----
ws_reg = wb.create_sheet("Regional Breakdown")
reg_data = defaultdict(lambda: {"units": 0, "revenue": 0.0})
for row in SALES_DATA:
    reg = row[2]
    units = row[3]
    rev = units * row[4]
    reg_data[reg]["units"] += units
    reg_data[reg]["revenue"] += rev

for c, h in enumerate(["Region", "Units Sold", "Revenue"], 1):
    cell = ws_reg.cell(row=1, column=c, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, hdr_align, thin_border

for r, (reg, vals) in enumerate(sorted(reg_data.items()), 2):
    ws_reg.cell(row=r, column=1, value=reg)
    ws_reg.cell(row=r, column=2, value=vals["units"])
    ws_reg.cell(row=r, column=3, value=round(vals["revenue"], 2))
    for c in range(1, 4):
        cell = ws_reg.cell(row=r, column=c)
        cell.font = Font(name="Calibri", size=10)
        cell.border = thin_border
        if c >= 2:
            cell.alignment = Alignment(horizontal="right")
    ws_reg.cell(row=r, column=3).number_format = curr_fmt

# Pie chart
pie = PieChart()
pie.title = "Revenue Share by Region"
pie.width, pie.height = 18, 14
pie.add_data(Reference(ws_reg, min_col=3, min_row=1, max_row=len(reg_data) + 1),
             titles_from_data=True)
pie.set_categories(Reference(ws_reg, min_col=1, min_row=2, max_row=len(reg_data) + 1))
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
for i in range(len(reg_data)):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = chart_colors[i % 6]
    pie.series[0].data_points.append(pt)
ws_reg.add_chart(pie, "E2")
for c, w in [('A', 14), ('B', 14), ('C', 14)]:
    ws_reg.column_dimensions[c].width = w

# ----- SHEET 4: Executive Dashboard -----
ws_exec = wb.create_sheet("Executive Dashboard")
ws_exec.merge_cells("A1:F1")
title = ws_exec.cell(row=1, column=1, value="SALES PERFORMANCE DASHBOARD — Q1 2025")
title.font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
title.alignment = Alignment(horizontal="center", vertical="center")
ws_exec.row_dimensions[1].height = 35

total_revenue = sum(r[3] * r[4] for r in SALES_DATA)
total_units = sum(r[3] for r in SALES_DATA)
total_profit = sum((r[3] * r[4]) - (r[3] * r[5]) for r in SALES_DATA)
avg_margin = total_profit / total_revenue if total_revenue else 0

kpis = [("TOTAL REVENUE", f"${total_revenue:,.0f}", 1),
        ("TOTAL UNITS SOLD", f"{total_units:,}", 2),
        ("TOTAL PROFIT", f"${total_profit:,.0f}", 3),
        ("AVG MARGIN", f"{avg_margin:.1%}", 4)]
for kpi_title, kpi_val, col in kpis:
    for r_offset, (val, font) in enumerate(
        [(kpi_title, Font(size=10, bold=True, color=WHITE)),
         (kpi_val, Font(size=22, bold=True, color=WHITE))]):
        cell = ws_exec.cell(row=3 + r_offset, column=col, value=val)
        cell.font = font
        cell.fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    ws_exec.column_dimensions[get_column_letter(col)].width = 22
ws_exec.row_dimensions[4].height = 40

# Summary table on dashboard
ws_exec.merge_cells("A6:D6")
ws_exec.cell(row=6, column=1, value="Key Metrics by Category").font = Font(
    size=12, bold=True, color=DARK_BLUE)
for c, h in enumerate(["Category", "Units", "Revenue", "Profit"], 1):
    cell = ws_exec.cell(row=7, column=c, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, hdr_align, thin_border
for r, (cat, vals) in enumerate(sorted_cats, 8):
    ws_exec.cell(row=r, column=1, value=cat)
    ws_exec.cell(row=r, column=2, value=vals["units"])
    ws_exec.cell(row=r, column=3, value=round(vals["revenue"], 2))
    ws_exec.cell(row=r, column=4, value=round(vals["profit"], 2))
    for c in range(1, 5):
        cell = ws_exec.cell(row=r, column=c)
        cell.font = Font(name="Calibri", size=10)
        cell.border = thin_border
        if r % 2 == 0:
            cell.fill = alt_fill
    ws_exec.cell(row=r, column=3).number_format = curr_fmt
    ws_exec.cell(row=r, column=4).number_format = curr_fmt

# ============================================================
# SAVE
# ============================================================
output_path = "Sales_Report_Q1_2025.xlsx"
wb.save(output_path)
print(f"✅ Report generated: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Total Revenue: ${total_revenue:,.2f}")
print(f"   Total Profit: ${total_profit:,.2f}")
print(f"   Avg Margin: {avg_margin:.1%}")
